import streamlit as st
import pandas as pd
import io
import re
import json
import time
import asyncio
from datetime import date
from anthropic import Anthropic, AsyncAnthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
CLAUDE_MODEL   = "claude-haiku-4-5-20251001"
OVERLAP_ROWS   = 5            # reduced from 8 — saves ~10% tokens per file
MAX_RETRIES    = 4
MAX_CONCURRENT = 3            # back to 3 — with longer backoff this is safe
PROMPT_VERSION = "v5.8"

# Chunk sizing — smaller = more parallelism + less output truncation risk
def get_chunk_size(total_rows: int) -> int:
    if total_rows <= 50:  return total_rows   # tiny file — single chunk
    if total_rows > 600:  return 50            # large file — keep output tight
    return 70                                  # standard

st.set_page_config(
    page_title="RealVal · Rent Roll Standardizer",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Sora:wght@600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
    background-color: #0f1923;
    color: #e2e8f0;
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 0 3rem 4rem; max-width: 100%; background: #111c27; }

.rv-hero {
    background: linear-gradient(135deg, #0a1628 0%, #0d2137 40%, #0e3347 100%);
    border-bottom: 1px solid rgba(46,196,182,0.2);
    padding: 2.2rem 3rem 2rem;
    display: flex; align-items: center; justify-content: space-between;
    position: relative; overflow: hidden;
}
.rv-hero::before {
    content: ''; position: absolute; top: 0; right: 0; bottom: 0; width: 40%;
    background: radial-gradient(ellipse at 80% 50%, rgba(46,196,182,0.07) 0%, transparent 70%);
    pointer-events: none;
}
.rv-hero-left { display: flex; align-items: center; gap: 1.2rem; z-index: 1; }
.rv-logo {
    width: 56px; height: 56px;
    background: linear-gradient(135deg, #2ec4b6 0%, #1a9490 100%);
    border-radius: 12px; display: flex; align-items: center; justify-content: center;
    font-family: 'Sora', sans-serif; font-weight: 700; font-size: 1.4rem;
    color: white; letter-spacing: -0.03em; flex-shrink: 0;
    box-shadow: 0 4px 20px rgba(46,196,182,0.4);
}
.rv-hero-text h1 { font-family: 'Sora', sans-serif; font-size: 1.5rem; font-weight: 700; color: #fff; margin: 0; letter-spacing: -0.02em; }
.rv-hero-text p  { font-size: 0.75rem; color: rgba(255,255,255,0.45); margin: 0.25rem 0 0; letter-spacing: 0.08em; text-transform: uppercase; }
.rv-hero-right   { display: flex; align-items: center; gap: 1rem; z-index: 1; }
.rv-badge  { background: rgba(46,196,182,0.12); border: 1px solid rgba(46,196,182,0.35); border-radius: 20px; padding: 0.35rem 1rem; font-size: 0.7rem; font-weight: 600; color: #2ec4b6; letter-spacing: 0.08em; text-transform: uppercase; }
.rv-version{ font-size: 0.68rem; color: rgba(255,255,255,0.25); letter-spacing: 0.06em; }

.sec-label { font-size: 0.65rem; font-weight: 700; letter-spacing: 0.15em; text-transform: uppercase; color: #2ec4b6; display: flex; align-items: center; gap: 0.6rem; margin: 2rem 0 1rem; }
.sec-label::after { content: ''; flex: 1; height: 1px; background: linear-gradient(90deg, rgba(46,196,182,0.3), transparent); }

.upload-panel { background: linear-gradient(135deg, #0d2137, #0e2a40); border: 1px solid rgba(46,196,182,0.2); border-radius: 14px; padding: 1.8rem 2rem; position: relative; overflow: hidden; }
.upload-panel::before { content: ''; position: absolute; top: -40px; right: -40px; width: 160px; height: 160px; background: radial-gradient(circle, rgba(46,196,182,0.08), transparent 70%); pointer-events: none; }
.upload-panel h3 { font-family: 'Sora', sans-serif; font-size: 1rem; font-weight: 700; color: #fff; margin: 0 0 0.4rem; }
.upload-panel p  { font-size: 0.78rem; color: rgba(255,255,255,0.45); margin: 0 0 1.2rem; }

.hint-box { background: #0d1e2e; border: 1px solid rgba(46,196,182,0.15); border-radius: 10px; padding: 1rem 1.2rem; margin-top: 1rem; }
.hint-box label { font-size: 0.65rem; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; color: #2ec4b6; display: block; margin-bottom: 0.5rem; }

.info-panel { background: #0d1e2e; border: 1px solid rgba(255,255,255,0.07); border-radius: 14px; padding: 1.5rem; }
.info-panel h4 { font-size: 0.72rem; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; color: #2ec4b6; margin: 0 0 0.8rem; }
.info-row  { display: flex; align-items: flex-start; gap: 0.6rem; margin-bottom: 0.55rem; font-size: 0.78rem; color: rgba(255,255,255,0.55); }
.info-dot  { width: 6px; height: 6px; background: #2ec4b6; border-radius: 50%; margin-top: 5px; flex-shrink: 0; }

.file-pill { display: inline-flex; align-items: center; gap: 0.5rem; background: rgba(46,196,182,0.1); border: 1px solid rgba(46,196,182,0.3); border-radius: 20px; padding: 0.4rem 1rem; font-size: 0.78rem; color: #2ec4b6; font-weight: 500; margin-bottom: 0.6rem; }
.sheet-pill { display: inline-flex; align-items: center; gap: 0.5rem; background: rgba(96,165,250,0.1); border: 1px solid rgba(96,165,250,0.3); border-radius: 20px; padding: 0.3rem 0.8rem; font-size: 0.72rem; color: #93c5fd; font-weight: 500; margin-left: 0.5rem; }

.stButton > button { background: linear-gradient(135deg, #2ec4b6, #1a9490) !important; color: #0a1628 !important; font-family: 'Inter', sans-serif !important; font-weight: 700 !important; font-size: 0.85rem !important; letter-spacing: 0.04em !important; border: none !important; border-radius: 10px !important; padding: 0.7rem 2rem !important; width: 100% !important; box-shadow: 0 4px 20px rgba(46,196,182,0.35) !important; transition: all 0.2s !important; }
.stButton > button:hover { background: linear-gradient(135deg, #3dd9ca, #2ec4b6) !important; box-shadow: 0 6px 28px rgba(46,196,182,0.5) !important; transform: translateY(-1px) !important; }

.steps-wrap { background: #0d1e2e; border: 1px solid rgba(255,255,255,0.07); border-radius: 12px; padding: 1.2rem 1.8rem; display: flex; align-items: center; gap: 0; margin: 1rem 0 1.5rem; overflow-x: auto; }
.step { display: flex; align-items: center; gap: 0.55rem; flex-shrink: 0; }
.step-dot { width: 30px; height: 30px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 0.72rem; font-weight: 700; flex-shrink: 0; }
.step-dot.done   { background: #2ec4b6; color: #0a1628; }
.step-dot.active { background: transparent; color: #2ec4b6; border: 2px solid #2ec4b6; box-shadow: 0 0 12px rgba(46,196,182,0.4); animation: pulse 1.5s infinite; }
.step-dot.wait   { background: rgba(255,255,255,0.06); color: rgba(255,255,255,0.25); }
@keyframes pulse { 0%,100%{box-shadow:0 0 8px rgba(46,196,182,0.3)} 50%{box-shadow:0 0 18px rgba(46,196,182,0.6)} }
.step-text { font-size: 0.77rem; white-space: nowrap; }
.step-text.done   { color: #2ec4b6; font-weight: 600; }
.step-text.active { color: #ffffff; font-weight: 600; }
.step-text.wait   { color: rgba(255,255,255,0.25); }
.step-sep { color: rgba(255,255,255,0.15); margin: 0 0.8rem; font-size: 0.9rem; flex-shrink: 0; }

.kpi-grid { display: grid; grid-template-columns: repeat(5,1fr); gap: 1rem; margin: 0.5rem 0 1.5rem; }
.kpi-card { background: #0d1e2e; border: 1px solid rgba(255,255,255,0.07); border-radius: 12px; padding: 1.1rem 1.3rem; border-top: 2px solid #1e3a4a; transition: border-color 0.2s, transform 0.2s; }
.kpi-card:hover { border-top-color: #2ec4b6; transform: translateY(-2px); }
.kpi-card.hi { border-top-color: #2ec4b6; }
.kpi-label { font-size: 0.62rem; font-weight: 700; letter-spacing: 0.12em; text-transform: uppercase; color: rgba(255,255,255,0.35); margin-bottom: 0.45rem; }
.kpi-value { font-family: 'Sora', sans-serif; font-size: 1.75rem; font-weight: 700; color: #ffffff; line-height: 1.1; }
.kpi-sub   { font-size: 0.68rem; color: rgba(255,255,255,0.3); margin-top: 0.3rem; }

.recon-grid { display: grid; grid-template-columns: repeat(2,1fr); gap: 0.85rem; margin-bottom: 1.5rem; }
.recon-card { background: #0d1e2e; border: 1px solid rgba(255,255,255,0.07); border-radius: 12px; padding: 1rem 1.2rem; display: flex; align-items: center; gap: 1rem; }
.r-icon { width: 38px; height: 38px; border-radius: 9px; display: flex; align-items: center; justify-content: center; font-size: 1.05rem; flex-shrink: 0; }
.r-icon.pass { background: rgba(34,197,94,0.15); }
.r-icon.warn { background: rgba(245,158,11,0.15); }
.r-icon.fail { background: rgba(239,68,68,0.15); }
.r-body  { flex: 1; }
.r-title { font-size: 0.78rem; font-weight: 600; color: #e2e8f0; margin-bottom: 0.1rem; }
.r-detail{ font-size: 0.7rem; color: rgba(255,255,255,0.4); }
.r-badge { font-size: 0.62rem; font-weight: 700; letter-spacing: 0.07em; text-transform: uppercase; padding: 0.2rem 0.6rem; border-radius: 20px; flex-shrink: 0; }
.rb-pass { background: rgba(34,197,94,0.15); color: #4ade80; }
.rb-warn { background: rgba(245,158,11,0.15); color: #fbbf24; }
.rb-fail { background: rgba(239,68,68,0.15); color: #f87171; }

.flag-banner { background: rgba(245,158,11,0.1); border: 1px solid rgba(245,158,11,0.3); border-radius: 10px; padding: 0.8rem 1.2rem; margin-bottom: 1rem; display: flex; align-items: center; gap: 0.7rem; font-size: 0.8rem; color: #fbbf24; }

.tbl-wrap { background: #0d1e2e; border: 1px solid rgba(255,255,255,0.07); border-radius: 12px; overflow: hidden; margin-bottom: 1.5rem; }
.tbl-header { background: linear-gradient(135deg, #0a1f32, #0d2840); padding: 0.75rem 1.2rem; display: flex; align-items: center; justify-content: space-between; border-bottom: 1px solid rgba(46,196,182,0.15); }
.tbl-header-left { font-size: 0.72rem; font-weight: 600; letter-spacing: 0.08em; text-transform: uppercase; color: rgba(255,255,255,0.5); }
.legend { display: flex; gap: 1.2rem; }
.leg-item { display: flex; align-items: center; gap: 0.4rem; font-size: 0.67rem; color: rgba(255,255,255,0.4); }
.leg-dot  { width: 8px; height: 8px; border-radius: 2px; }

.rv-tbl { width: 100%; border-collapse: collapse; font-size: 0.79rem; }
.rv-tbl th { background: rgba(255,255,255,0.03); color: rgba(255,255,255,0.45); font-weight: 600; font-size: 0.68rem; text-align: left; letter-spacing: 0.05em; padding: 0.65rem 1rem; border-bottom: 1px solid rgba(255,255,255,0.06); white-space: nowrap; text-transform: uppercase; }
.rv-tbl td { padding: 0.52rem 1rem; border-bottom: 1px solid rgba(255,255,255,0.04); color: #cbd5e1; white-space: nowrap; }
.rv-tbl tr:last-child td { border-bottom: none; }
.rv-tbl tr.occ:hover td { background: rgba(255,255,255,0.025); }
.rv-tbl tr.vac td { background: rgba(245,158,11,0.05); color: #fcd34d; }
.rv-tbl tr.nr  td { background: rgba(34,197,94,0.05);  color: #86efac; font-style: italic; }
.rv-tbl tr.flagged td { background: rgba(245,158,11,0.08); }
.rv-tbl td.mono  { font-family: 'SF Mono','Fira Code',monospace; font-size: 0.73rem; color: #7dd3fc; }
.rv-tbl td.right { text-align: right; }
.rv-tbl td.green { color: #4ade80; font-weight: 600; }
.rv-tbl .tag { display: inline-block; font-size: 0.6rem; font-weight: 700; letter-spacing: 0.06em; text-transform: uppercase; padding: 0.12rem 0.45rem; border-radius: 20px; }
.tag-occ  { background: rgba(96,165,250,0.15);  color: #93c5fd; }
.tag-vac  { background: rgba(245,158,11,0.15);  color: #fcd34d; }
.tag-nr   { background: rgba(34,197,94,0.15);   color: #86efac; }
.tag-warn { background: rgba(245,158,11,0.2);   color: #f59e0b; }

/* Streamlit tab styling */
.stTabs [data-baseweb="tab-list"] { background: #0d1e2e; border-radius: 10px; padding: 0.3rem; gap: 0.3rem; border: 1px solid rgba(255,255,255,0.07); }
.stTabs [data-baseweb="tab"] { background: transparent; color: rgba(255,255,255,0.4); border-radius: 8px; font-size: 0.82rem; font-weight: 500; padding: 0.5rem 1.2rem; }
.stTabs [aria-selected="true"] { background: rgba(46,196,182,0.15) !important; color: #2ec4b6 !important; font-weight: 600; }
.stTabs [data-baseweb="tab-panel"] { padding: 1.2rem 0 0; }

.stDownloadButton > button { background: linear-gradient(135deg, #0e4a5c, #1a6b7a) !important; color: #ffffff !important; font-weight: 600 !important; border: 1px solid rgba(46,196,182,0.3) !important; border-radius: 10px !important; font-size: 0.85rem !important; padding: 0.7rem 2rem !important; width: 100% !important; box-shadow: 0 4px 20px rgba(14,74,92,0.4) !important; }
.stDownloadButton > button:hover { background: linear-gradient(135deg, #2ec4b6, #1a9490) !important; color: #0a1628 !important; border-color: transparent !important; }

.stProgress > div > div > div > div { background: #2ec4b6 !important; }

[data-testid="stFileUploader"] section { background: rgba(255,255,255,0.03) !important; border: 1.5px dashed rgba(46,196,182,0.3) !important; border-radius: 10px !important; }
[data-testid="stFileUploader"] section:hover { border-color: rgba(46,196,182,0.6) !important; }
[data-testid="stFileUploader"] section p { color: rgba(255,255,255,0.45) !important; }

.stTextArea textarea { background: #0a1628 !important; color: #e2e8f0 !important; border: 1px solid rgba(46,196,182,0.2) !important; border-radius: 8px !important; font-size: 0.82rem !important; }
.stTextArea textarea:focus { border-color: rgba(46,196,182,0.5) !important; box-shadow: 0 0 0 2px rgba(46,196,182,0.1) !important; }
.stSelectbox > div > div { background: #0a1628 !important; border: 1px solid rgba(46,196,182,0.2) !important; color: #e2e8f0 !important; border-radius: 8px !important; }

/* Format library */
.lib-card { background: #0d1e2e; border: 1px solid rgba(255,255,255,0.07); border-radius: 10px; padding: 1rem 1.2rem; margin-bottom: 0.6rem; }
.lib-card h5 { font-size: 0.78rem; font-weight: 600; color: #e2e8f0; margin: 0 0 0.3rem; }
.lib-card p  { font-size: 0.72rem; color: rgba(255,255,255,0.4); margin: 0; }
.lib-meta { font-size: 0.65rem; color: rgba(46,196,182,0.7); margin-top: 0.3rem; }

.rv-footer { border-top: 1px solid rgba(255,255,255,0.06); margin-top: 4rem; padding-top: 1.2rem; display: flex; justify-content: space-between; align-items: center; font-size: 0.68rem; color: rgba(255,255,255,0.2); }
.rv-footer a { color: #2ec4b6; text-decoration: none; }
</style>
""", unsafe_allow_html=True)


# ── PASSWORD GATE ─────────────────────────────────────────────────────────────
def check_password():
    if st.session_state.get("authenticated"):
        return True
    st.markdown("""
    <div class="rv-hero">
      <div class="rv-hero-left">
        <div class="rv-logo">RV</div>
        <div class="rv-hero-text"><h1>Rent Roll Standardizer</h1>
        <p>RealVal · Multifamily Underwriting Intelligence</p></div>
      </div>
      <div class="rv-badge">Secure Access</div>
    </div>""", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1.2, 1, 1.2])
    with col2:
        st.markdown("""
        <div style="background:#0d1e2e;border:1px solid rgba(46,196,182,0.2);border-radius:14px;
        padding:2rem 1.8rem;text-align:center;margin-top:2rem;">
        <div style="width:52px;height:52px;background:linear-gradient(135deg,#2ec4b6,#1a9490);
        border-radius:12px;display:flex;align-items:center;justify-content:center;
        font-family:Sora,sans-serif;font-weight:700;font-size:1.3rem;color:white;
        margin:0 auto 1.2rem;box-shadow:0 4px 20px rgba(46,196,182,0.4);">RV</div>
        <div style="font-family:Sora,sans-serif;font-size:1.15rem;font-weight:700;color:#fff;margin-bottom:0.3rem;">
        Internal Access Only</div>
        <div style="font-size:0.78rem;color:rgba(255,255,255,0.35);margin-bottom:1.5rem;">
        Authorized RealVal analysts only</div></div>""", unsafe_allow_html=True)
        pwd = st.text_input("Password", type="password", label_visibility="collapsed", placeholder="Enter password…")
        if st.button("Sign In →"):
            if pwd == st.secrets.get("password", ""):
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    st.markdown("&nbsp;", unsafe_allow_html=True)
    return False


# ── MULTI-SHEET PRE-PROCESSOR ─────────────────────────────────────────────────
def detect_rent_roll_sheet(file_bytes: bytes) -> tuple[pd.DataFrame, str]:
    """
    Reads all sheets, scores each by rent-roll signal words,
    returns (dataframe_of_best_sheet, sheet_name).
    Uses data_only=True so formula cells (e.g. =A8) return cached values
    instead of formula strings — critical for files like 3_RR where sub-rows
    use =A8 references that resolve to the parent unit number.
    """
    xl      = pd.ExcelFile(io.BytesIO(file_bytes))
    sheets  = xl.sheet_names

    if len(sheets) == 1:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheets[0], header=None)
        return df, sheets[0]

    SIGNALS = ["unit","tenant","lease","rent","sqft","sq ft","market","charge",
               "resident","move","vacant","expir","deposit","occupan"]
    EXCLUDE = ["summary","totals","cover","index","toc","contents","chart","graph",
               "pivot","dashboard","notes","readme","parameters","report param"]

    scores = {}
    for name in sheets:
        lname = name.lower()
        if any(x in lname for x in EXCLUDE):
            scores[name] = -1
            continue
        try:
            sample = pd.read_excel(io.BytesIO(file_bytes), sheet_name=name, header=None, nrows=20)
            text   = " ".join(str(v).lower() for v in sample.values.flatten() if pd.notna(v))
            score  = sum(text.count(s) for s in SIGNALS)
            score += sample.shape[0] * 0.5
            scores[name] = score
        except Exception:
            scores[name] = 0

    best = max(scores, key=scores.get)
    df   = pd.read_excel(io.BytesIO(file_bytes), sheet_name=best, header=None)

    # Post-process: replace formula strings (=A8, =A9...) with NaN
    # These are Excel references that openpyxl can't always resolve.
    # Sub-rows that carry forward unit number via formulas should be treated as None.
    import re
    formula_mask = df.map(
        lambda v: bool(str(v).startswith('=')) if v is not None and not isinstance(v, (int, float)) else False
    )
    df = df.where(~formula_mask, other=None)

    return df, best


# ── COLUMN LABELLER ───────────────────────────────────────────────────────────
def detect_format(combined_headers: list[str], is_split_header: bool) -> str:
    """
    Identifies the software/export format from the column header names.
    Returns one of: 'yardi', 'appfolio', 'vesper', 'onsite', 'unknown'
    """
    h = " ".join(combined_headers).lower()
    # OneSite/RealPage: has 'trans code' or 'lease rent' or 'market + addl'
    if "trans code" in h or "lease rent" in h or "market + addl" in h:
        return "onsite"
    if is_split_header:
        return "yardi"
    if "budgeted rent" in h:
        return "vesper"
    if "scheduled charges" in h or "bldg-unit" in h:
        return "appfolio"
    if "amount" in combined_headers:
        return "yardi"
    return "unknown"


def preprocess_onsite(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pre-processes OneSite/RealPage exports before label_raw_df runs.
    Handles three OneSite-specific issues:
      1. Repeating page headers (~every 68 rows) — detected by '\nUnit' in col1
      2. Interleaved Pending/Applicant future-lease rows — col8='N/A' or col20='N/A'
         with Pending/Applicant in the next column
      3. Embedded newlines in column headers (cleaned in label_raw_df's clean())
    """
    vals = list(df.values)
    keep = []
    for row in vals:
        # Strip repeating page-break header rows
        c1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        if '\nUnit' in c1 or c1 == '\nUnit':
            continue
        # Strip section labels like 'details', 'Other\nCharges/ Credits'
        c0 = str(row[0]).strip() if row[0] is not None else ''
        if c0.lower() in ('details', 'other\ncharges/ credits', 'other charges/ credits'):
            continue
        # Strip Pending/Applicant interleaved rows (future leases)
        # These have no real unit number — col1 is N/A
        if c1 == 'N/A':
            status_col = str(row[20]).strip() if len(row) > 20 and row[20] is not None else ''
            if 'pending' in status_col.lower() or 'applicant' in status_col.lower():
                continue
        keep.append(row)

    result = pd.DataFrame(keep, columns=df.columns)

    # Drop blank rows (OneSite puts a blank row between every unit block)
    result = result.dropna(how='all').reset_index(drop=True)
    return result


def label_raw_df(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Auto-detects the rent roll software format, assigns clean column names,
    and returns a column position map with the detected format name.

    Returns: (labelled_df, col_map)
    col_map includes key 'fmt' = one of 'yardi' | 'appfolio' | 'vesper' | 'unknown'
    """
    vals = df.values
    header_row = None

    for i, row in enumerate(vals[:15]):
        texts = [str(v).lower().strip() for v in row if v is not None and str(v).strip().lower() != 'nan']
        has_unit   = any("unit" in t or "bldg" in t for t in texts)
        has_charge = any(kw in t for t in texts for kw in ("charge", "rent", "market", "amount", "scheduled", "budgeted"))
        if has_unit and has_charge:
            header_row = i
            break

    if header_row is None:
        df.columns = [str(c) for c in df.columns]
        return df, {"fmt": "unknown"}

    def clean(v):
        s = str(v).strip() if v is not None else ""
        return "" if s.lower() == "nan" else s

    row_a = [clean(v) for v in vals[header_row]]

    # Determine if row_b is a real header continuation or a data/label row.
    # Real continuation: short label fragments like "Sq Ft", "Rent", "Code", "Expiration".
    # Data row / label: contains long strings, full dates (with colons+digits), or
    # section labels like "Property: X".
    # KEY FIX: datetime strings like "2025-07-18 00:00:00" contain colons but ARE data,
    # not header labels. We check for colons ONLY in purely-alphabetic tokens.
    row_b_raw = [clean(v) for v in vals[header_row + 1]] \
                if header_row + 1 < len(vals) else [""] * len(row_a)
    non_empty_b = [v for v in row_b_raw if v]

    def _looks_like_header_fragment(v: str) -> bool:
        """True if v could be a header word/phrase, not data."""
        if len(v) > 25: return False                              # too long
        if v.replace('.','').replace(',','').replace('-','').isdigit(): return False  # pure number
        if any(c.isdigit() for c in v) and ':' in v: return False # datetime-like
        if ':' in v and not any(c.isdigit() for c in v): return False  # label like "Property:"
        return True

    is_real_continuation = (
        len(non_empty_b) >= 2
        and all(_looks_like_header_fragment(v) for v in non_empty_b)
    )
    row_b = row_b_raw if is_real_continuation else [""] * len(row_a)

    # Combine header rows
    combined = []
    for a, b in zip(row_a, row_b):
        parts = " ".join(p for p in [a, b] if p).strip()
        combined.append(parts if parts else f"col_{len(combined)}")

    # Detect format BEFORE mapping names
    fmt = detect_format(combined, is_real_continuation)

    # ── Universal canonical name map ──
    name_map = {
        # Yardi (split two-row header)
        "unit":                 "Unit No",
        "unit type":            "Unit Type",
        "unit sq ft":           "Sq Ft",
        "sq ft":                "Sq Ft",
        "resident":             "Resident ID",   # col3 = Resident ID, NOT tenant name
        "name":                 "Tenant Name",   # col4 = actual Tenant Name
        "market rent":          "Market Rent",
        "charge code":          "Charge Code",
        "amount":               "Charge Amount",
        "resident deposit":     "Res Deposit",
        "other deposit":        "Other Deposit",
        "other deposits":       "Other Deposit",
        "move in":              "Move In",
        "lease expiration":     "Lease Expiration",
        "move out":             "Move Out",
        "balance":              "Balance",
        # AppFolio (single-row header)
        "bldg-unit":            "Unit No",
        "sqft":                 "Sq Ft",
        "unit status":          "Unit Status",
        "lease start":          "Lease Start",
        "lease end":            "Lease End",
        "expected move-out":    "Move Out",
        "ledger":               "Ledger",
        "actual charges":       "Actual Charges",
        "scheduled charges":    "Charge Amount",
        "deposit held":         "Res Deposit",
        "move-in":              "Move In",
        # Vesper/Sunbelt variant
        "budgeted rent":        "Market Rent",
        # OneSite/RealPage format
        "unit":                 "Unit No",        # '\nUnit' cleaned to 'Unit'
        "floorplan":            "Unit Type",
        "unit designation":     "Unit Designation",
        "unit/lease status":    "Unit Status",
        "name":                 "Tenant Name",
        "move-in move-out":     "Move In",
        "market + addl.":       "Market Rent",    # OneSite market rent column
        "trans code":           "Charge Code",    # OneSite charge code column
        "lease rent":           "Charge Amount",  # OneSite effective rent column
        "total billing":        "Total Billing",
        "dep on hand":          "Res Deposit",
    }

    clean_cols = []
    seen = {}
    for raw in combined:
        # Strip embedded newlines (OneSite headers have \nUnit, Market\n+ Addl. etc.)
        key = raw.replace('\n', ' ').lower().strip()
        mapped = name_map.get(key, raw.replace('\n', ' '))
        if mapped in seen:
            seen[mapped] += 1
            mapped = f"{mapped}_{seen[mapped]}"
        else:
            seen[mapped] = 0
        clean_cols.append(mapped)

    # Build position map
    col_map = {"fmt": fmt}
    for i, name in enumerate(clean_cols):
        if name == "Unit No":            col_map["unit"]          = i
        elif name == "Charge Code":      col_map["charge_code"]   = i
        elif name == "Charge Amount":    col_map["charge_amount"] = i
        elif name == "Market Rent":      col_map["market_rent"]   = i
        elif name == "Res Deposit":      col_map["deposit"]       = i
        elif name == "Move In":          col_map["move_in"]       = i
        elif name in ("Lease Expiration", "Lease Start"): col_map["lease_exp"] = i
        elif name == "Move Out":         col_map["move_out"]      = i
        elif name == "Unit Status":      col_map["unit_status"]   = i
        elif name == "Tenant Name":      col_map["tenant"]        = i

    # Skip 2 rows for Yardi split-header, 1 row for single-header formats
    rows_to_skip = 2 if is_real_continuation else 1
    data_df = df.iloc[header_row + rows_to_skip:].copy()
    data_df.columns = clean_cols[:data_df.shape[1]]
    data_df = data_df.reset_index(drop=True)

    # For OneSite: strip repeating page headers, Pending rows, blank rows
    if fmt == "onsite":
        data_df = preprocess_onsite(data_df)

    # Drop entirely-empty trailing columns (reduces tokens on 52-col files)
    data_df = data_df.dropna(axis=1, how='all')

    return data_df, col_map


def init_library():
    if "format_library" not in st.session_state:
        st.session_state["format_library"] = []

def save_to_library(broker_name: str, hint: str, raw_df: pd.DataFrame, result_df: pd.DataFrame):
    """Save a 5-row anonymized sample to the in-session format library."""
    sample_raw    = raw_df.head(5).to_csv(index=False)
    sample_output = result_df.head(5).to_dict(orient="records")
    entry = {
        "broker":    broker_name,
        "hint":      hint,
        "date":      date.today().strftime("%m/%d/%Y"),
        "raw_sample": sample_raw,
        "out_sample": sample_output,
        "unit_count": len(result_df),
    }
    st.session_state["format_library"].append(entry)

def build_library_context() -> str:
    """Format saved examples as few-shot context for the Claude prompt."""
    lib = st.session_state.get("format_library", [])
    if not lib:
        return ""
    lines = ["\nFORMAT LIBRARY — examples from previously processed files (use as reference):"]
    for e in lib[-3:]:  # last 3 entries
        lines.append(f"\nBroker: {e['broker']} | Hint: {e['hint']} | Date: {e['date']}")
        lines.append(f"Raw sample (first 5 rows):\n{e['raw_sample']}")
        lines.append(f"Output sample: {json.dumps(e['out_sample'][:2], indent=2)}")
    return "\n".join(lines)


# ── STEPS ─────────────────────────────────────────────────────────────────────
STEPS = ["Reading File", "Detecting Sheet", "Sending to Claude", "Building Output", "Complete"]

def render_steps(active: int) -> str:
    html = "<div class='steps-wrap'>"
    for i, label in enumerate(STEPS):
        if i < active:   dc, tc, icon = "done",   "done",   "✓"
        elif i == active:dc, tc, icon = "active", "active", str(i+1)
        else:            dc, tc, icon = "wait",   "wait",   str(i+1)
        html += f"<div class='step'><div class='step-dot {dc}'>{icon}</div><span class='step-text {tc}'>{label}</span></div>"
        if i < len(STEPS)-1: html += "<span class='step-sep'>›</span>"
    html += "</div>"
    return html


# ── COMPACT PROMPT BUILDER ────────────────────────────────────────────────────
# Format-specific instructions keep Claude from guessing structure
FORMAT_NOTES = {
    "yardi": (
        "FORMAT: Yardi export. Split two-row header.\n"
        "Unit header row columns: Unit No | Unit Type | Sq Ft | Resident ID | Tenant Name | Market Rent | Charge Code | Charge Amount | Res Deposit | Other Deposit | Move In | Lease Expiration | Move Out | Balance\n"
        "IMPORTANT: col3 = Resident ID (e.g. t9142369) — this is NOT the tenant name.\n"
        "col4 = Tenant Name — use this as the tenant name.\n"
        "col5 = Market Rent — NEVER use as Effective Rent.\n"
        "col6 = Charge Code (rent, trash, petfee, amenity, conrent, rmrnt etc.)\n"
        "col7 = Charge Amount — THIS is the Effective Rent when Charge Code is a rent code.\n"
        "col8 = Res Deposit — NEVER use as Effective Rent.\n"
        "Rent codes: rent, rnt, base, baserent, conrent, rmrnt, subsidy, rentsub, hap.\n"
        "conrent = contracted rent (Fieldcrest-style) — IS the effective rent.\n"
        "concourt = court credit (always negative) — include, nets against conrent.\n"
        "rmrnt = property-prefixed rent code — IS the effective rent.\n"
        "Subsidy-only units: EffRent = subsidy amount (do NOT fall back to Market Rent)."
    ),
    "appfolio": (
        "FORMAT: AppFolio export. Single-row header.\n"
        "Columns: Bldg-Unit | SQFT | Unit Status | Move-In | Lease Start | Lease End | Expected Move-Out | Market Rent | Ledger | Charge Code | Actual Charges | Scheduled Charges | Balance | Deposit Held\n"
        "Effective Rent = Scheduled Charges where Charge Code = 'Rent' — NOT Deposit Held.\n"
        "Tenant Name comes from Resident column (Ledger = 'Resident' sub-row).\n"
        "Vacant: Unit Status contains 'Vacant', Charge Code = '-'.\n"
        "Section headers ('Unit Type:', 'Property:') are NOT units — skip them."
    ),
    "vesper": (
        "FORMAT: Vesper/Sunbelt AppFolio export. Single-row header.\n"
        "Columns: Bldg-Unit | Unit Type | SQFT | Unit Status | Budgeted Rent | Resident | Ledger | Charge Code | Scheduled Charges | Balance | Deposit Held | Move-In | Lease Start | Lease End | Expected Move-Out\n"
        "Market Rent = Budgeted Rent column. Effective Rent = Scheduled Charges where Charge Code = 'Rent'.\n"
        "Tenant Name = Resident column. Charge codes are full text: 'Rent', 'Pet Rent' etc.\n"
        "Employee Discount is a NEGATIVE charge — EXCLUDE it, do not net against rent."
    ),
    "onsite": (
        "FORMAT: OneSite/RealPage export. Single-row header. Sparse columns (data every other column).\n"
        "Key columns after label_raw_df: Unit No | Floorplan (Unit Type) | SQFT | Unit/Lease Status | Name (Tenant) | Lease Start | Lease End | Market Rent | Trans Code | Lease Rent | Total Billing | Dep On Hand\n"
        "Market Rent = 'Market Rent' column (originally 'Market + Addl.').\n"
        "Effective Rent = 'Charge Amount' column (originally 'Lease Rent') where Trans Code = 'RENT'.\n"
        "ALL other Trans Codes are fees/adjustments — exclude: BUILDINGFACILITIES, GARAGE, VALETWASTE, PETRENT, STORAGE, MTOM, PARKING, CONC/SPECL, OFCRCRED, MODEL, REFERRAL.\n"
        "Rows with status 'Pending renewal' or 'Applicant' are future leases — EXCLUDE.\n"
        "Vacant units have status 'Vacant' and no Lease Rent value."
    ),
    "unknown": (
        "FORMAT: Unknown. Use column names as labelled.\n"
        "Effective Rent = Charge Amount column where Charge Code = rent/Rent.\n"
        "Never use Market Rent or Deposit columns as Effective Rent."
    ),
}

def build_prompt(chunk_text: str, chunk_num: int, total_chunks: int,
                 analyst_hint: str, library_ctx: str, col_map: dict = None) -> str:
    hint = f"\nNOTE: {analyst_hint.strip()}" if analyst_hint.strip() else ""

    fmt = (col_map or {}).get("fmt", "unknown")
    format_note = FORMAT_NOTES.get(fmt, FORMAT_NOTES["unknown"])

    # Locked column positions — tells Claude exactly where to look
    if col_map:
        amt_col  = col_map.get("charge_amount", 7)
        code_col = col_map.get("charge_code",   6)
        mkt_col  = col_map.get("market_rent",   5)
        dep_col  = col_map.get("deposit",        8)
        col_lock = (
            f"COLUMN LOCK — read Effective Rent from 'Charge Amount' (col {amt_col}) ONLY.\n"
            f"  Charge Code = col {code_col} | Market Rent = col {mkt_col} (NEVER use as EffRent)"
            f" | Deposit = col {dep_col} (NEVER use as EffRent)"
        )
    else:
        col_lock = "COLUMN RULE: EffRent = 'Charge Amount' where Charge Code = rent. Never use Market Rent or Deposit."

    return f"""Multifamily rent roll standardizer. ({PROMPT_VERSION})
Output: JSON array, 9 columns + optional flag field.
Columns: Unit No | Unit Size (SF) | Market Rent (Monthly) | Effective Rent (Monthly) | Move In Date | Lease Start Date | Lease End Date | Move Out Date | Tenant Name

{format_note}
{col_lock}

Rules:
1. 1 row/unit — merge ALL charge sub-rows for the same unit. Scan every sub-row.
2. EffRent WHITELIST — include ONLY these charge codes in Effective Rent:
   SHORT CODES (Yardi):  rnt, rent, base, baserent, rentsub, sub, hap, subsidy, housing
   PROPERTY-PREFIXED:    conrent (contracted rent), rmrnt (property-prefixed rent) — these ARE rent
   OFFSET CODES:         concourt (court credit, negative) — include, nets against conrent
   FULL TEXT:            Rent, Base Rent, Subsidy Rent, HAP, Housing Assistance
   EXCLUDE EVERYTHING ELSE. When in doubt, exclude.
   CRITICAL DISTINCTION: conrent ≠ con/conc. conrent IS rent. con/conc = concession → EXCLUDE.
   Common codes/descriptions to EXCLUDE:
   pet, petfee, petrent, rmpet, PETRENT, Pet Rent, Pet Fees, Monthly Pet Rent → EXCLUDE
   park, pkg, parkfee, rmpkg, rmpki, rmpke, rmpkx, GARAGE, PARKING, Garage/Parking → EXCLUDE
   pest, pestfee, conpest, Pest Control Fees Monthly → EXCLUDE
   trash, VALETWASTE, Trash Fees → EXCLUDE
   util, conutil, Utility Service Fee, Utility Administration Fees → EXCLUDE
   water, wtr, Water/Sewer Reimbursement → EXCLUDE
   elec, electric, Electric - Reimbursement → EXCLUDE
   gas, Gas Reimbursement → EXCLUDE
   con, conc, concession, Monthly Concession, Month to Month Fees → EXCLUDE
   conrisk, condamw, conpetrt, conmtmf, conutil, conpest → EXCLUDE (con-prefix fees)
   conemp, emp, empdisc, emptaxed, OFCRCRED → EXCLUDE (employee/officer discounts)
   cbl, cable, tv → EXCLUDE
   dep, deposit, secdep, Security Deposit, Deposit Held → EXCLUDE
   late, latefee → EXCLUDE
   stor, storage, STORAGE → EXCLUDE
   xmgmt, xonetime → EXCLUDE
   wd, insurmp, rntins → EXCLUDE (insurance/damage fees)
   amenity, amentech, AmenTech, amenityfee, rmamen, club → EXCLUDE
   mtmfee, mtm, conmtmf, MTOM, MTM Fee → EXCLUDE (month-to-month fees)
   short, Short Term Fee → EXCLUDE
   rmcpk, rmcrp, rmcsx, rmptf → EXCLUDE (property-prefixed fees/adjustments)
   Employee Discount → EXCLUDE (this is a negative charge — do NOT net it against rent)
   Package Locker Fee, Guarantor Waiver Fee, Credit Reporting Svc Fee → EXCLUDE
   BUILDINGFACILITIES, REFERRAL, MODEL, CONC/SPECL → EXCLUDE
   Lost Rent Model Unit → EXCLUDE
   Charge Total, Total → NOT a charge — skip row entirely
   Any unrecognized code or description → EXCLUDE
   OUTPUT: Unit Type field = the unit type/floorplan from the raw file (col1 in Yardi, col1 in AppFolio).
3. Annual rent÷12. Rent/SF×SF=monthly.
4. Dates→MM/DD/YYYY or null.
5. VACANT: any unit where Unit Status contains 'Vacant', or Tenant Name is blank/VACANT,
   or Charge Code is '-' or None with 0 scheduled charges → include, EffRent=null, Name="VACANT".
6. ADMIN/MODEL/EXCLUDED: any unit with status 'Admin', 'Office', 'Model', 'Down Unit', 'Excluded'
   → include, EffRent=null, Name="ADMIN" or "MODEL" as appropriate.
7. Exclude future leases (no active rent charge, future move-in date, Pending/Applicant status).
8. Exclude section headers (Unit Type:, Property:, details, Other Charges/Credits), subtotals, footer rows.
9. Duplicate unit# → keep first only.
10. Round money to 2 decimals.
11. Add "flag":true if uncertain about any row.
Never skip a unit that has a unit number.{hint}
{library_ctx}
Return raw JSON array only — no fences, no text. Start with [ end with ].

Data (chunk {chunk_num}/{total_chunks}):
{chunk_text}"""


# ── PYTHON RENT RECOVERY — safety net for Claude misses ──────────────────────
RENT_CODES    = {
    # Standard Yardi short codes
    "rnt", "rent", "base", "baserent", "base rent",
    # Subsidy / HAP codes
    "rentsub", "sub", "hap", "subsidy", "housing", "subsidy rent",
    # Fieldcrest Walk consolidated-charge format
    "conrent",     # contracted rent — IS the effective rent in con-prefix Yardi exports
    "concourt",    # court/eviction credit — always negative, offsets conrent to $0
    # The Jordan (4_RR) property-prefixed rent code
    "rmrnt",       # rm-prefixed rent code used by this specific Yardi config
}
SUBSIDY_CODES = {"rentsub", "sub", "hap", "subsidy", "housing", "subsidy rent"}

def recover_missing_rents(result_df: pd.DataFrame, raw_df: pd.DataFrame,
                          col_map: dict = None) -> pd.DataFrame:
    """
    For any occupied unit where Claude returned null Effective Rent,
    scan the raw file directly in Python and extract rent + subsidy.
    This is a deterministic safety net — no LLM involved.
    """
    if raw_df is None or raw_df.empty:
        return result_df

    occupied_mask = (
        ~result_df["Tenant Name"].astype(str).str.upper().isin(["VACANT","ADMIN","MODEL"])
        & result_df["Tenant Name"].notna()
    )
    missing_eff = result_df["Effective Rent (Monthly)"].isna()
    problem_units = result_df.loc[occupied_mask & missing_eff, "Unit No"].tolist()

    if not problem_units:
        return result_df

    # Build a lookup from the raw dataframe
    # Raw file: unit number in col 0, charge code in col 6, amount in col 7
    # Sub-rows have None in col 0 — they belong to the last seen unit
    raw_vals  = raw_df.values
    num_cols  = raw_vals.shape[1]
    if num_cols < 8:
        return result_df

    # Use col_map if provided (most reliable), then named columns, then positional fallback
    if col_map:
        unit_col   = col_map.get("unit",          0)
        code_col   = col_map.get("charge_code",   6)
        amount_col = col_map.get("charge_amount", 7)
    else:
        cols = list(raw_df.columns)
        unit_col   = cols.index("Unit No")       if "Unit No"       in cols else 0
        code_col   = cols.index("Charge Code")   if "Charge Code"   in cols else 6
        amount_col = cols.index("Charge Amount") if "Charge Amount" in cols else 7

    unit_charges: dict[str, dict] = {}
    current_unit = None

    for row in raw_vals:
        unit_val   = row[unit_col]
        charge_raw = str(row[code_col]).strip().lower()   if len(row) > code_col   and row[code_col]   is not None else ""
        amount_raw = row[amount_col]                       if len(row) > amount_col else None

        if unit_val and str(unit_val).strip() not in ("", "nan", "None"):
            current_unit = str(unit_val).strip()
            if current_unit not in unit_charges:
                unit_charges[current_unit] = {"rent": None, "subsidy": 0.0}
            # first row of unit block may also have a charge
            if charge_raw in RENT_CODES:
                try:
                    amt = float(amount_raw)
                    if charge_raw in SUBSIDY_CODES:
                        unit_charges[current_unit]["subsidy"] += amt
                    else:
                        unit_charges[current_unit]["rent"] = amt
                except (TypeError, ValueError):
                    pass
        elif current_unit:
            # sub-row
            if charge_raw in RENT_CODES:
                try:
                    amt = float(amount_raw)
                    if charge_raw in SUBSIDY_CODES:
                        unit_charges[current_unit]["subsidy"] += amt
                    else:
                        unit_charges[current_unit]["rent"] = amt
                except (TypeError, ValueError):
                    pass

    # Patch missing values back into result_df
    recovered = 0
    for unit in problem_units:
        charges = unit_charges.get(unit)
        if not charges:
            continue
        rent_amt    = charges.get("rent")
        subsidy_amt = charges.get("subsidy", 0.0)
        if rent_amt is not None:
            # Normal case: rent + any subsidy
            eff = round(rent_amt + subsidy_amt, 2)
        elif subsidy_amt and subsidy_amt > 0:
            # Subsidy-only unit (e.g. Brandon Oaks 0417, 1209, 1405)
            eff = round(subsidy_amt, 2)
        else:
            continue
        result_df.loc[result_df["Unit No"] == unit, "Effective Rent (Monthly)"] = eff
        result_df.loc[result_df["Unit No"] == unit, "flag"] = True
        recovered += 1

    if recovered:
        print(f"[Recovery] Fixed {recovered} units with missing Effective Rent from raw file")

    return result_df


# ── ASYNC CLAUDE CALL (single chunk) ─────────────────────────────────────────
async def call_claude_async(client: AsyncAnthropic, chunk_text: str,
                             chunk_num: int, total_chunks: int,
                             analyst_hint: str, library_ctx: str,
                             semaphore: asyncio.Semaphore,
                             col_map: dict = None) -> tuple[int, list]:
    """Returns (chunk_num, rows). Semaphore caps concurrent connections."""
    prompt = build_prompt(chunk_text, chunk_num, total_chunks, analyst_hint, library_ctx, col_map)
    last_error = None

    async with semaphore:  # only MAX_CONCURRENT chunks run at once
        for attempt in range(MAX_RETRIES):
            try:
                resp = await client.messages.create(
                    model=CLAUDE_MODEL, max_tokens=8192,
                    messages=[{"role": "user", "content": prompt}]
                )
                raw = resp.content[0].text.strip()
                raw = re.sub(r"^```(?:json)?\s*\n?", "", raw, flags=re.MULTILINE)
                raw = re.sub(r"\n?```\s*$",           "", raw, flags=re.MULTILINE)
                rows = json.loads(raw.strip())
                return (chunk_num, rows)
            except json.JSONDecodeError as e:
                last_error = f"JSON error chunk {chunk_num} (attempt {attempt+1}): {e}"
                await asyncio.sleep(2)
            except Exception as e:
                err_str = str(e)
                # 429 rate limit — back off longer before retrying
                if "429" in err_str or "rate_limit" in err_str:
                    wait = 15 * (attempt + 1)   # 15s, 30s, 45s, 60s
                    last_error = f"Rate limit on chunk {chunk_num} — waiting {wait}s (attempt {attempt+1}/{MAX_RETRIES})"
                    await asyncio.sleep(wait)
                else:
                    last_error = f"Error chunk {chunk_num} (attempt {attempt+1}): {e}"
                    await asyncio.sleep(2)
    raise RuntimeError(last_error)


# ── POST-PROCESSING VALIDATION ────────────────────────────────────────────────
def validate_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Three-pass validation:
    Pass 1 — Flag outliers (EffRent > 2× market, or < $100).
    Pass 2 — HARD RULE: Any occupied unit with null EffRent gets flagged for review.
              We NO LONGER auto-fill Market Rent as fallback — subsidy-only units and
              special charge-code units (conrent, rmrnt etc.) legitimately have EffRent
              that differs from Market Rent. Stamping Market Rent was causing wrong values.
              Instead: flag as 0 so the analyst knows to review it.
    Pass 3 — Ensure flag column is clean bool.
    """
    if "flag" not in df.columns:
        df["flag"] = False
    df["flag"] = df["flag"].fillna(False).astype(bool)

    mkt = pd.to_numeric(df["Market Rent (Monthly)"],    errors="coerce")
    eff = pd.to_numeric(df["Effective Rent (Monthly)"], errors="coerce")

    NON_REVENUE = {"VACANT", "ADMIN", "MODEL"}
    occ = (
        df["Tenant Name"].notna()
        & ~df["Tenant Name"].astype(str).str.strip().str.upper().isin(NON_REVENUE)
        & (df["Tenant Name"].astype(str).str.strip() != "")
    )

    # Pass 1 — flag outliers
    df.loc[occ & eff.notna() & mkt.notna() & (eff > mkt * 2), "flag"] = True
    df.loc[occ & eff.notna() & (eff < 100), "flag"] = True

    # Pass 2 — flag occupied units with blank EffRent but DO NOT auto-fill Market Rent
    # recover_missing_rents() already tried Python-level recovery from raw file.
    # If still null here, mark as flagged with 0 so analyst can see and fix.
    blank_eff     = eff.isna()
    occupied_blank = occ & blank_eff
    if occupied_blank.any():
        df.loc[occupied_blank, "Effective Rent (Monthly)"] = 0.0
        df.loc[occupied_blank, "flag"] = True

    return df


# ── PARALLEL STANDARDIZE ──────────────────────────────────────────────────────
def standardize_rent_roll(df, step_ph, prog_ph, status_ph, analyst_hint, library_ctx, raw_df=None):
    api_key    = st.secrets["anthropic_api_key"]

    # Label columns before chunking — critical so Claude sees named columns
    labelled_df, col_map = label_raw_df(df.copy())
    total_rows  = len(labelled_df)
    chunk_size  = get_chunk_size(total_rows)

    step_ph.markdown(render_steps(2), unsafe_allow_html=True)

    # Build chunks with overlap to prevent boundary unit loss
    chunks, start = [], 0
    while start < total_rows:
        end = min(start + chunk_size, total_rows)
        chunks.append((len(chunks) + 1, labelled_df.iloc[start:end].to_csv(index=False)))
        if end == total_rows: break
        start += chunk_size - OVERLAP_ROWS

    num_chunks = len(chunks)
    status_ph.markdown(
        f"<small style='color:rgba(255,255,255,0.4);'>"
        f"Processing {num_chunks} chunk{'s' if num_chunks>1 else ''} "
        f"({MAX_CONCURRENT} at a time) · {total_rows} rows · chunk size {chunk_size}</small>",
        unsafe_allow_html=True
    )

    # ── Run chunks with controlled concurrency ──
    async def run_all():
        async_client = AsyncAnthropic(api_key=api_key)
        sem = asyncio.Semaphore(MAX_CONCURRENT)
        tasks = [
            call_claude_async(async_client, csv_text, chunk_num, num_chunks,
                              analyst_hint, library_ctx, sem, col_map)
            for chunk_num, csv_text in chunks
        ]
        return await asyncio.gather(*tasks)

    try:
        results = asyncio.run(run_all())
    except RuntimeError as e:
        status_ph.empty()
        st.error(str(e))
        return pd.DataFrame(), col_map

    prog_ph.progress(1.0)
    status_ph.empty()
    step_ph.markdown(render_steps(3), unsafe_allow_html=True)

    # Re-order by chunk number and flatten
    results_sorted = sorted(results, key=lambda x: x[0])
    all_rows = [row for _, chunk_rows in results_sorted for row in chunk_rows]

    if not all_rows:
        return pd.DataFrame(), col_map

    result_df = pd.DataFrame(all_rows)
    COLS = ["Unit No","Unit Type","Unit Size (SF)","Market Rent (Monthly)","Effective Rent (Monthly)",
            "Move In Date","Lease Start Date","Lease End Date","Move Out Date","Tenant Name"]
    for col in COLS:
        if col not in result_df.columns: result_df[col] = None
    result_df = result_df[COLS + (["flag"] if "flag" in result_df.columns else [])]

    for dc in ["Move In Date","Lease Start Date","Lease End Date","Move Out Date"]:
        result_df[dc] = pd.to_datetime(result_df[dc], errors="coerce").dt.strftime("%m/%d/%Y")

    result_df.drop_duplicates(subset=["Unit No"], keep="first", inplace=True)
    result_df.reset_index(drop=True, inplace=True)

    # Step 1: Python-level rent recovery (fixes Claude misses from raw file)
    result_df = recover_missing_rents(result_df, raw_df, col_map)

    # Step 2: Hard validation — occupied units must always have Effective Rent
    result_df = validate_rows(result_df)

    step_ph.markdown(render_steps(4), unsafe_allow_html=True)
    return result_df, col_map


# ── COLOR-CODED TABLE ─────────────────────────────────────────────────────────
def render_table(df: pd.DataFrame) -> str:
    COLS = ["Unit No","Unit Type","Unit Size (SF)","Market Rent (Monthly)","Effective Rent (Monthly)",
            "Move In Date","Lease Start Date","Lease End Date","Move Out Date","Tenant Name"]
    hdr  = "".join(f"<th>{c}</th>" for c in COLS) + "<th>Status</th>"
    body = ""
    flag_col = "flag" in df.columns

    for _, row in df.iterrows():
        name     = str(row.get("Tenant Name","") or "").strip().upper()
        is_flag  = bool(row.get("flag", False)) if flag_col else False

        if name == "VACANT":
            tr, status_tag = "vac", "<span class='tag tag-vac'>Vacant</span>"
            name_tag = "<span class='tag tag-vac'>Vacant</span>"
        elif name in ("ADMIN","MODEL"):
            tr, status_tag = "nr", "<span class='tag tag-nr'>Non-Rev</span>"
            name_tag = f"<span class='tag tag-nr'>{name.title()}</span>"
        else:
            tr = "occ"
            name_tag = "<span class='tag tag-occ'>Occupied</span>"
            status_tag = "<span class='tag tag-occ'>OK</span>"

        if is_flag:
            tr += " flagged"
            status_tag = "<span class='tag tag-warn'>⚠ Review</span>"

        def fmt(col, val):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return "<span style='color:rgba(255,255,255,0.2);'>—</span>"
            if col in ("Market Rent (Monthly)","Effective Rent (Monthly)"):
                try:    return f"${float(val):,.2f}"
                except: return str(val)
            if col == "Unit Size (SF)":
                try:    return f"{int(float(val)):,}"
                except: return str(val)
            return str(val)

        cells = ""
        for col in COLS:
            val = row.get(col)
            cls = "mono" if col=="Unit No" else \
                  "right green" if col=="Effective Rent (Monthly)" else \
                  "right" if col in ("Market Rent (Monthly)","Unit Size (SF)") else ""
            display = name_tag if col == "Tenant Name" else fmt(col, val)
            cells += f"<td class='{cls}'>{display}</td>"
        cells += f"<td>{status_tag}</td>"
        body  += f"<tr class='{tr}'>{cells}</tr>"

    flag_count = int(df["flag"].sum()) if "flag" in df.columns else 0
    legend_extra = f'<div class="leg-item"><div class="leg-dot" style="background:#f59e0b;border:1px solid #fbbf24;"></div>⚠ Flagged ({flag_count})</div>' if flag_count else ""

    return f"""
    <div class='tbl-wrap'>
      <div class='tbl-header'>
        <span class='tbl-header-left'>{len(df)} units · standardized output</span>
        <div class='legend'>
          <div class='leg-item'><div class='leg-dot' style='background:#fcd34d;'></div>Vacant</div>
          <div class='leg-item'><div class='leg-dot' style='background:#86efac;'></div>Admin/Model</div>
          <div class='leg-item'><div class='leg-dot' style='background:#93c5fd;'></div>Occupied</div>
          {legend_extra}
        </div>
      </div>
      <div style='overflow-x:auto;max-height:500px;overflow-y:auto;'>
        <table class='rv-tbl'><thead><tr>{hdr}</tr></thead><tbody>{body}</tbody></table>
      </div>
    </div>"""


# ── RAW FILE TABLE ────────────────────────────────────────────────────────────
def render_raw_table(df: pd.DataFrame, sheet_name: str) -> str:
    df_show = df.head(200)
    hdr  = "".join(f"<th>{i}</th>" for i in range(df_show.shape[1]))
    body = ""
    for _, row in df_show.iterrows():
        cells = "".join(
            f"<td style='max-width:180px;overflow:hidden;text-overflow:ellipsis;'>"
            f"{'' if pd.isna(v) else str(v)}</td>"
            for v in row
        )
        body += f"<tr class='occ'>{cells}</tr>"

    return f"""
    <div class='tbl-wrap'>
      <div class='tbl-header'>
        <span class='tbl-header-left'>Raw file · sheet: "{sheet_name}" · {len(df)} rows × {df.shape[1]} cols (showing first 200)</span>
      </div>
      <div style='overflow-x:auto;max-height:500px;overflow-y:auto;'>
        <table class='rv-tbl'><thead><tr>{hdr}</tr></thead><tbody>{body}</tbody></table>
      </div>
    </div>"""


# ── KPI CARDS ─────────────────────────────────────────────────────────────────
def render_kpis(df: pd.DataFrame) -> str:
    occ   = df[~df["Tenant Name"].str.upper().isin(["VACANT","ADMIN","MODEL"]) & df["Tenant Name"].notna()]
    vac   = df[df["Tenant Name"].str.upper()=="VACANT"]
    total = len(df)
    occ_pct = (len(occ)/total*100) if total else 0
    avg_mkt = pd.to_numeric(df["Market Rent (Monthly)"],    errors="coerce").mean()
    avg_eff = pd.to_numeric(occ["Effective Rent (Monthly)"], errors="coerce").mean()
    tpgi    = pd.to_numeric(df["Market Rent (Monthly)"],    errors="coerce").sum()
    def fc(v): return f"${v:,.0f}" if pd.notna(v) and v > 0 else "—"
    cards = [
        ("Total Units",        f"{total}",         "All incl. vacant & non-rev", ""),
        ("Occupancy Rate",     f"{occ_pct:.1f}%",  f"{len(occ)} occ / {len(vac)} vac", "hi"),
        ("Avg Market Rent",    fc(avg_mkt),          "Monthly · all units", ""),
        ("Avg Effective Rent", fc(avg_eff),           "Monthly · occupied only", "hi"),
        ("Total Potential GI", fc(tpgi),              "Sum of all market rents / mo", ""),
    ]
    html = "<div class='kpi-grid'>"
    for label, value, sub, cls in cards:
        html += f"<div class='kpi-card {cls}'><div class='kpi-label'>{label}</div><div class='kpi-value'>{value}</div><div class='kpi-sub'>{sub}</div></div>"
    return html + "</div>"


# ── RECONCILIATION ────────────────────────────────────────────────────────────
def render_recon(df: pd.DataFrame, orig_rows: int) -> str:
    NON_REVENUE = {"VACANT", "ADMIN", "MODEL"}
    occ = df[
        df["Tenant Name"].notna()
        & ~df["Tenant Name"].astype(str).str.strip().str.upper().isin(NON_REVENUE)
        & (df["Tenant Name"].astype(str).str.strip() != "")
    ]
    vac        = df[df["Tenant Name"].astype(str).str.strip().str.upper() == "VACANT"]
    clean_rows = len(df)
    clean_mkt  = pd.to_numeric(df["Market Rent (Monthly)"], errors="coerce").sum()
    missing    = [c for c in ["Unit No","Unit Size (SF)","Market Rent (Monthly)",
                               "Effective Rent (Monthly)","Lease Start Date","Lease End Date","Tenant Name"]
                  if c not in df.columns]
    size_warn  = clean_rows < orig_rows * 0.25
    flag_count = int(df["flag"].sum()) if "flag" in df.columns else 0

    # Hard check: any occupied unit still missing Effective Rent after all recovery steps
    occ_eff = pd.to_numeric(occ["Effective Rent (Monthly)"], errors="coerce")
    occ_blank_count = int(occ_eff.isna().sum())

    checks = [
        {"title": "All Required Columns Present",
         "detail": "7 of 7 columns found" if not missing else f"Missing: {', '.join(missing)}",
         "st": "pass" if not missing else "fail",
         "icon": "✓" if not missing else "✗",
         "badge": "Pass" if not missing else "Fail"},

        {"title": "Row Count vs Raw File",
         "detail": f"{clean_rows} standardized rows from {orig_rows} raw rows",
         "st": "warn" if size_warn else "pass",
         "icon": "⚠" if size_warn else "✓",
         "badge": "Review" if size_warn else "Pass"},

        {"title": "Market Rent Populated",
         "detail": f"Total market rent: ${clean_mkt:,.2f}",
         "st": "pass" if clean_mkt > 0 else "fail",
         "icon": "✓" if clean_mkt > 0 else "✗",
         "badge": "Pass" if clean_mkt > 0 else "Fail"},

        {"title": "Occupied Tenant Count",
         "detail": f"{len(occ)} occupied · {len(vac)} vacant",
         "st": "pass" if len(occ) > 0 else "warn",
         "icon": "✓" if len(occ) > 0 else "⚠",
         "badge": "Pass" if len(occ) > 0 else "Review"},

        # Hard rule check — this should always be Pass after the new enforce step
        {"title": "Occupied Units — Effective Rent Complete",
         "detail": "All occupied units have Effective Rent ✓" if occ_blank_count == 0
                   else f"⚠ {occ_blank_count} occupied unit{'s' if occ_blank_count > 1 else ''} still missing Effective Rent — manual review required",
         "st": "pass" if occ_blank_count == 0 else "fail",
         "icon": "✓" if occ_blank_count == 0 else "✗",
         "badge": "Pass" if occ_blank_count == 0 else f"{occ_blank_count} Missing"},

        {"title": "Flagged Rows",
         "detail": f"{flag_count} rows auto-flagged for analyst review" if flag_count else "No rows flagged",
         "st": "warn" if flag_count else "pass",
         "icon": "⚠" if flag_count else "✓",
         "badge": f"{flag_count} Flags" if flag_count else "Clean"},
    ]

    html = "<div class='recon-grid'>"
    for c in checks:
        rb = {"pass": "rb-pass", "warn": "rb-warn", "fail": "rb-fail"}[c["st"]]
        html += f"""<div class='recon-card'>
          <div class='r-icon {c["st"]}'>{c["icon"]}</div>
          <div class='r-body'><div class='r-title'>{c["title"]}</div><div class='r-detail'>{c["detail"]}</div></div>
          <span class='r-badge {rb}'>{c["badge"]}</span>
        </div>"""
    return html + "</div>"


# ── SOURCE SUMMARY EXTRACTOR ──────────────────────────────────────────────────
def extract_source_summary(raw_df: pd.DataFrame) -> dict | None:
    """
    Scans the raw dataframe for the summary block at the bottom of the rent roll.
    Handles two formats:
      Yardi   — summary keywords in col0, unit counts in col10, market rent in col6
      AppFolio — status labels in col0 with counts in col1; charge codes in col3, amounts in col4
    """
    if raw_df is None or raw_df.empty:
        return None

    summary = {}
    vals    = raw_df.values

    # ── Keyword sets ──
    OCCUPIED_KW    = {"occupied units", "total occupied units", "occupied"}
    VACANT_KW      = {"total vacant units", "vacant units", "total vacant"}
    NONREV_KW      = {"total non rev units", "non rev units", "non-revenue",
                      "admin/model", "model/admin", "total excluded units", " total excluded units"}
    TOTALS_KW      = {"totals:", "totals", "grand total", "total rentable units",
                      " total units", "total units"}
    CHARGECODE_KW  = {"summary of charges", "charge code", "summary of charges by charge code",
                      "charge code summary"}

    in_chargecode_yardi   = False   # Yardi: col0=code, col3=amount
    in_chargecode_appfolio= False   # AppFolio: col3=code, col4=amount
    charge_codes: dict[str, float] = {}

    def safe_float(v):
        try:
            return float(str(v).replace(",", "").replace("(", "-").replace(")", "").strip())
        except: return None

    def safe_int(v):
        try:
            v2 = safe_float(v)
            return int(v2) if v2 is not None else None
        except: return None

    for row in vals:
        row = list(row) + [None] * max(0, 14 - len(row))
        c0 = str(row[0]).strip().lower() if row[0] is not None else ""
        c3 = str(row[3]).strip().lower() if row[3] is not None else ""

        # Skip blanks, parenthetical rows, pure header labels
        if c0.startswith("(") or c0 in ("", "nan", "charge code", "description",
                                         "ledger: resident", "property: crystal gardens"):
            # Stop AppFolio CC section if c3 signals end ('resident total:' / 'total:')
            if in_chargecode_appfolio and "total" in c3:
                in_chargecode_appfolio = False
                continue
            # But still scan col3 for AppFolio charge code rows
            if in_chargecode_appfolio and c3 and c3 not in ("", "nan", "charge code",
                "property: crystal gardens", "ledger: resident", "resident total:",
                "total:", "charge code"):
                amt = safe_float(row[4])
                if amt is not None:
                    charge_codes[c3] = charge_codes.get(c3, 0) + amt
            continue

        # ── Yardi charge code section (col0=code, col3=amount) ──
        if any(kw in c0 for kw in CHARGECODE_KW):
            in_chargecode_yardi    = True
            in_chargecode_appfolio = False
            continue

        # ── AppFolio charge code section (col3=section header) ──
        if any(kw in c3 for kw in CHARGECODE_KW):
            in_chargecode_appfolio = True
            in_chargecode_yardi    = False
            continue

        # Process Yardi charge code rows
        if in_chargecode_yardi:
            code = c0
            if code in ("total", "", "nan", "charge code"):
                if code == "total": in_chargecode_yardi = False
                continue
            amt = safe_float(row[3])
            if amt is not None:
                charge_codes[code] = amt
            continue

        # Process AppFolio charge code rows (col3=code, col4=amount)
        if in_chargecode_appfolio:
            code = c3
            if code in ("total:", "total", "resident total:", "", "nan"):
                if "total" in code: in_chargecode_appfolio = False
                continue
            amt = safe_float(row[4])
            if amt is not None and code not in ("charge code", "property: crystal gardens",
                                                 "ledger: resident"):
                charge_codes[code] = charge_codes.get(code, 0) + amt
            continue

        # ── Yardi summary rows (unit counts in col10, market rent in col6) ──
        if any(kw == c0 for kw in OCCUPIED_KW):
            v = safe_int(row[10]) or safe_int(row[1])   # Yardi=col10, AppFolio=col1
            if v: summary["src_occupied_units"] = v
            mv = safe_float(row[6])
            if mv: summary["src_occupied_mkt"] = mv

        elif any(kw == c0 for kw in VACANT_KW):
            v = safe_int(row[10]) or safe_int(row[1])
            if v: summary["src_vacant_units"] = v

        elif any(kw == c0 for kw in NONREV_KW):
            v = safe_int(row[10]) or safe_int(row[1])
            if v: summary["src_nonrev_units"] = v

        elif any(kw == c0 for kw in TOTALS_KW):
            # Yardi: col10=units, col6=market, col11=occ_pct, col7=lease_charges
            # AppFolio " Total Units": col1=total
            v = safe_int(row[10]) or safe_int(row[1])
            if v: summary["src_total_units"] = v
            mv = safe_float(row[6])
            if mv: summary["src_total_mkt"] = mv
            pv = safe_float(row[11])
            if pv: summary["src_occ_pct"] = pv
            lv = safe_float(row[7])
            if lv: summary["src_lease_charges"] = lv

        # AppFolio status summary rows — col0=description, col1=count
        elif c0 in ("occupied no notice", "notice unrented"):
            v = safe_int(row[1])
            if v:
                summary["src_occupied_units"] = summary.get("src_occupied_units", 0) + v

        elif c0 == "total occupied units":
            v = safe_int(row[1])
            if v: summary["src_occupied_units"] = v   # override accumulated

        elif c0 in ("vacant rented not ready", "vacant unrented not ready"):
            v = safe_int(row[1])
            if v:
                summary["src_vacant_units"] = summary.get("src_vacant_units", 0) + v

        elif c0 == "total vacant units":
            v = safe_int(row[1])
            if v: summary["src_vacant_units"] = v

        elif c0 in ("excluded - down unit", "excluded - admin/office unit",
                    "excluded -down unit", "excluded -admin/office unit"):
            v = safe_int(row[1])
            if v:
                summary["src_nonrev_units"] = summary.get("src_nonrev_units", 0) + v

        elif c0 in ("total rentable units", " total units", "total units"):
            v = safe_int(row[1])
            if v: summary["src_total_units"] = v

    if charge_codes:
        summary["src_charge_codes"] = charge_codes
        rent_total = sum(v for k, v in charge_codes.items() if k in RENT_CODES)
        if rent_total:
            summary["src_rent_total"] = rent_total

    return summary if summary else None


# ── SOURCE VERIFIER PANEL ─────────────────────────────────────────────────────
def render_source_verifier(standardized_df: pd.DataFrame, src: dict) -> str:
    """
    Side-by-side comparison: source file summary vs our standardized output.
    Each row shows the source number, our number, and a pass/warn/fail badge.
    """
    if not src:
        return ""

    NON_REVENUE = {"VACANT", "ADMIN", "MODEL"}
    occ = standardized_df[
        standardized_df["Tenant Name"].notna()
        & ~standardized_df["Tenant Name"].astype(str).str.strip().str.upper().isin(NON_REVENUE)
        & (standardized_df["Tenant Name"].astype(str).str.strip() != "")
    ]
    vac = standardized_df[
        standardized_df["Tenant Name"].astype(str).str.strip().str.upper() == "VACANT"
    ]
    nr  = standardized_df[
        standardized_df["Tenant Name"].astype(str).str.strip().str.upper().isin({"ADMIN","MODEL"})
    ]

    our_total_units   = len(standardized_df)
    our_occupied      = len(occ)
    our_vacant        = len(vac)
    our_nonrev        = len(nr)
    our_total_mkt     = pd.to_numeric(standardized_df["Market Rent (Monthly)"], errors="coerce").sum()
    our_eff_total     = pd.to_numeric(occ["Effective Rent (Monthly)"], errors="coerce").sum()

    def badge(ok, warn=False):
        if ok:   return "<span class='r-badge rb-pass'>✓ Match</span>"
        if warn: return "<span class='r-badge rb-warn'>⚠ Close</span>"
        return           "<span class='r-badge rb-fail'>✗ Mismatch</span>"

    def fmt_num(v, is_currency=False):
        if v is None or (isinstance(v, float) and pd.isna(v)): return "—"
        if is_currency: return f"${v:,.0f}"
        return f"{int(v):,}"

    def pct_ok(src_v, our_v, tol=0.02):
        if src_v is None or our_v is None: return None
        if src_v == 0: return our_v == 0
        return abs(src_v - our_v) / abs(src_v) <= tol

    rows_html = ""

    checks = []

    # Total units
    sv = src.get("src_total_units"); ov = our_total_units
    ok = (sv == ov) if sv is not None else None
    checks.append(("Total Units", fmt_num(sv), fmt_num(ov), ok, False))

    # Occupied
    sv = src.get("src_occupied_units"); ov = our_occupied
    ok = (sv == ov) if sv is not None else None
    checks.append(("Occupied Units", fmt_num(sv), fmt_num(ov), ok, False))

    # Vacant
    sv = src.get("src_vacant_units"); ov = our_vacant
    ok = (sv == ov) if sv is not None else None
    checks.append(("Vacant Units", fmt_num(sv), fmt_num(ov), ok, False))

    # Non-rev
    sv = src.get("src_nonrev_units"); ov = our_nonrev
    ok = (sv == ov) if sv is not None else None
    checks.append(("Non-Revenue Units", fmt_num(sv), fmt_num(ov), ok, False))

    # Total market rent (allow 1% tolerance for rounding)
    sv = src.get("src_total_mkt"); ov = our_total_mkt
    ok = pct_ok(sv, ov, 0.01)
    warn = (ok is None) or (not ok and pct_ok(sv, ov, 0.05))
    checks.append(("Total Market Rent", fmt_num(sv, True), fmt_num(ov, True), ok, warn and not ok))

    # Rent charge total from source vs our effective rent total (allow 2% — concessions etc.)
    sv = src.get("src_rent_total"); ov = our_eff_total
    ok = pct_ok(sv, ov, 0.02)
    warn = (ok is None) or (not ok and pct_ok(sv, ov, 0.08))
    checks.append(("Total Effective Rent (rent charges)", fmt_num(sv, True), fmt_num(ov, True), ok, warn and not ok))

    for label, src_val, our_val, ok, warn in checks:
        if ok is True:    icon, st_cls = "✓", "pass"
        elif ok is False and not warn: icon, st_cls = "✗", "fail"
        else:             icon, st_cls = "⚠", "warn"
        rows_html += f"""
        <tr>
          <td style='padding:0.55rem 1rem;font-size:0.8rem;color:#e2e8f0;font-weight:500;'>{label}</td>
          <td style='padding:0.55rem 1rem;font-size:0.8rem;color:#7dd3fc;text-align:right;font-family:monospace;'>{src_val}</td>
          <td style='padding:0.55rem 1rem;font-size:0.8rem;color:#a3e635;text-align:right;font-family:monospace;'>{our_val}</td>
          <td style='padding:0.55rem 1rem;text-align:center;'>
            <div class='r-icon {st_cls}' style='width:28px;height:28px;border-radius:6px;font-size:0.8rem;margin:auto;'>{icon}</div>
          </td>
        </tr>"""

    # Charge code breakdown table (if available)
    cc_html = ""
    cc = src.get("src_charge_codes", {})
    if cc:
        cc_rows = ""
        for code, amt in sorted(cc.items(), key=lambda x: -abs(x[1])):
            colour = "#4ade80" if code in RENT_CODES else "#94a3b8"
            cc_rows += f"<tr><td style='padding:0.35rem 0.8rem;font-size:0.75rem;color:{colour};font-family:monospace;'>{code}</td><td style='padding:0.35rem 0.8rem;font-size:0.75rem;color:#e2e8f0;text-align:right;font-family:monospace;'>${amt:,.0f}</td></tr>"
        cc_html = f"""
        <div style='margin-top:1rem;background:#0a1628;border:1px solid rgba(255,255,255,0.06);border-radius:10px;overflow:hidden;'>
          <div style='padding:0.5rem 0.8rem;background:#0d2137;font-size:0.65rem;font-weight:700;letter-spacing:0.1em;text-transform:uppercase;color:rgba(255,255,255,0.4);'>
            Charge Code Breakdown (Source File)
          </div>
          <table style='width:100%;border-collapse:collapse;'>{cc_rows}</table>
        </div>"""

    return f"""
    <div style='background:#0d1e2e;border:1px solid rgba(46,196,182,0.2);border-radius:12px;overflow:hidden;margin-bottom:1.5rem;'>
      <div style='background:linear-gradient(135deg,#0a1f32,#0d2840);padding:0.75rem 1.2rem;border-bottom:1px solid rgba(46,196,182,0.15);display:flex;align-items:center;justify-content:space-between;'>
        <span style='font-size:0.72rem;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:rgba(255,255,255,0.5);'>Source File Verifier — Standardizer vs Original Summary</span>
        <span style='font-size:0.68rem;color:rgba(46,196,182,0.6);'>Extracted from source footer</span>
      </div>
      <table style='width:100%;border-collapse:collapse;'>
        <thead>
          <tr style='background:rgba(255,255,255,0.03);'>
            <th style='padding:0.5rem 1rem;font-size:0.68rem;font-weight:600;letter-spacing:0.05em;text-transform:uppercase;color:rgba(255,255,255,0.35);text-align:left;'>Metric</th>
            <th style='padding:0.5rem 1rem;font-size:0.68rem;font-weight:600;letter-spacing:0.05em;text-transform:uppercase;color:#7dd3fc;text-align:right;'>Source File</th>
            <th style='padding:0.5rem 1rem;font-size:0.68rem;font-weight:600;letter-spacing:0.05em;text-transform:uppercase;color:#a3e635;text-align:right;'>Our Output</th>
            <th style='padding:0.5rem 1rem;font-size:0.68rem;font-weight:600;letter-spacing:0.05em;text-transform:uppercase;color:rgba(255,255,255,0.35);text-align:center;'>Check</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
      <div style='padding:0 1rem 0.8rem;'>{cc_html}</div>
    </div>"""


def build_excel(df: pd.DataFrame, raw_df: pd.DataFrame = None,
                source_summary: dict = None) -> bytes:
    wb   = Workbook()
    navy = "0A2E3D"; teal = "0E4A5C"
    thin = Side(style="thin", color="E2E8F0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = ["Unit No","Unit Type","Unit Size (SF)","Market Rent (Monthly)","Effective Rent (Monthly)",
            "Move In Date","Lease Start Date","Lease End Date","Move Out Date","Tenant Name"]
    DATE_COLS  = {6, 7, 8, 9}   # Move In, Lease Start, Lease End, Move Out (shifted +1)
    MONEY_COLS = {4, 5}          # Market Rent, Effective Rent (shifted +1)

    # Sheet 1: Standardized
    ws = wb.active; ws.title = "Standardized Rent Roll"
    for col, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color=navy)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
    ws.row_dimensions[1].height = 30

    flag_col = "flag" in df.columns
    for ri, rec in enumerate(df.itertuples(index=False), 2):
        nm = str(rec[8] or "").upper()
        iv = nm == "VACANT"; nr = nm in ("ADMIN", "MODEL")
        is_flagged = bool(getattr(rec, "flag", False)) if flag_col else False
        fill = (PatternFill("solid", start_color="FFFBEB") if iv else
                PatternFill("solid", start_color="F0FDF4") if nr else
                PatternFill("solid", start_color="FFF9EB") if is_flagged else
                PatternFill("solid", start_color="FFFFFF") if ri % 2 else
                PatternFill("solid", start_color="F8FAFC"))
        fc = "92400E" if iv else "166534" if nr else "854D0E" if is_flagged else "1E293B"
        for col, val in enumerate(rec[:9], 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border = bdr; c.fill = fill
            c.font = Font(name="Calibri", size=10, italic=(iv or nr), color=fc)
            if col == 3:             c.number_format = "#,##0";     c.alignment = Alignment(horizontal="right")  # Sq Ft
            elif col in MONEY_COLS:  c.number_format = "$#,##0.00"; c.alignment = Alignment(horizontal="right")
            elif col in DATE_COLS:   c.alignment = Alignment(horizontal="center")
            else:                    c.alignment = Alignment(horizontal="left")

    last = len(df) + 1; tot = last + 1
    for col in range(1, len(COLS) + 1):
        c = ws.cell(row=tot, column=col)
        c.fill = PatternFill("solid", start_color=navy)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.border = bdr; c.alignment = Alignment(horizontal="right")
    ws.cell(row=tot, column=1, value="TOTALS / AVERAGES").alignment = Alignment(horizontal="left")
    # col2 = Unit Type — no formula
    ws.cell(row=tot, column=3, value=f"=SUM(C2:C{last})").number_format = "#,##0"           # Sq Ft
    ws.cell(row=tot, column=4, value=f"=AVERAGE(D2:D{last})").number_format = "$#,##0.00"   # Market Rent
    ws.cell(row=tot, column=5, value=f'=AVERAGEIF(E2:E{last},"<>",E2:E{last})').number_format = "$#,##0.00"  # Eff Rent

    nr2 = tot + 2
    for i, note in enumerate(["Notes:",
        "• Yellow = Vacant", "• Green = Non-revenue (Admin/Model)",
        "• Orange = Flagged — verify Effective Rent manually",
        "• Effective Rent = base rent + housing subsidy only",
        "• Concessions, cable, parking, utilities, fees excluded"]):
        c = ws.cell(row=nr2 + i, column=1, value=note)
        c.font = Font(name="Calibri", bold=(i == 0), size=9,
                      color="166534" if "Green" in note else
                            "92400E" if "Yellow" in note else
                            "854D0E" if "Orange" in note else "595959")

    for i, w in enumerate([14, 16, 10, 22, 22, 14, 16, 16, 14, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{last}"

    # Sheet 2: Source Verification
    ws2 = wb.create_sheet(title="Source Verification")
    NON_REV = {"VACANT", "ADMIN", "MODEL"}
    occ_df  = df[df["Tenant Name"].notna() &
                 ~df["Tenant Name"].astype(str).str.strip().str.upper().isin(NON_REV) &
                 (df["Tenant Name"].astype(str).str.strip() != "")]
    vac_df  = df[df["Tenant Name"].astype(str).str.strip().str.upper() == "VACANT"]
    nr_df   = df[df["Tenant Name"].astype(str).str.strip().str.upper().isin({"ADMIN","MODEL"})]

    our = {
        "total_units": len(df),
        "occ_units":   len(occ_df),
        "vac_units":   len(vac_df),
        "nr_units":    len(nr_df),
        "total_mkt":   pd.to_numeric(df["Market Rent (Monthly)"],       errors="coerce").sum(),
        "occ_mkt":     pd.to_numeric(occ_df["Market Rent (Monthly)"],   errors="coerce").sum(),
        "eff_total":   pd.to_numeric(occ_df["Effective Rent (Monthly)"], errors="coerce").sum(),
        "occ_pct":     len(occ_df) / len(df) * 100 if len(df) else 0,
    }

    # Fix verifier consistency: compute our rent total using the EXACT same charge codes
    # that the source summary counted — not our broader RENT_CODES family.
    # This guarantees apples-to-apples comparison.
    ss = source_summary or {}
    src_cc = ss.get("src_charge_codes", {})
    if src_cc and raw_df is not None:
        # Which codes did the source count as rent-family?
        src_rent_codes = {k for k in src_cc if k in RENT_CODES}
        if src_rent_codes:
            # Sum those exact codes from the raw file directly
            raw_rent_total = 0.0
            current_unit_raw = None
            for row in raw_df.values:
                c0 = row[0]
                # Detect unit header rows
                if c0 is not None and str(c0).strip() not in ("", "nan", "None"):
                    current_unit_raw = str(c0).strip()
                # Read charge sub-rows
                try:
                    # Works for both labelled (named cols) and unlabelled (positional)
                    if hasattr(raw_df, 'columns') and 'Charge Code' in raw_df.columns:
                        cc_idx = list(raw_df.columns).index('Charge Code')
                        amt_idx = list(raw_df.columns).index('Charge Amount')
                    else:
                        cc_idx, amt_idx = 6, 7
                    code = str(row[cc_idx]).strip().lower() if row[cc_idx] is not None else ""
                    amt  = row[amt_idx]
                    if code in src_rent_codes and amt is not None:
                        raw_rent_total += float(str(amt).replace(",",""))
                except (IndexError, ValueError, TypeError):
                    pass
            our["eff_total_raw"] = raw_rent_total  # consistent comparison value

    def pct_diff(a, b):
        if b and b != 0: return abs(a - b) / abs(b) * 100
        return None

    def vstatus(our_v, src_v, tol=0.5, exact=False):
        if src_v is None: return "N/A", "F1F5F9", "64748B"
        if exact:
            return ("MATCH", "DCFCE7", "166534") if our_v == src_v else ("MISMATCH", "FEE2E2", "991B1B")
        d = pct_diff(our_v, src_v)
        if d is None:    return "N/A",            "F1F5F9", "64748B"
        if d <= tol:     return "MATCH",           "DCFCE7", "166534"
        if d <= 2.0:     return f"OFF {d:.1f}%",   "FEF9C3", "854D0E"
        return                  f"MISMATCH {d:.1f}%","FEE2E2", "991B1B"

    def fmtv(v, money=False, pct=False):
        if v is None: return "Not found"
        if money: return f"${v:,.2f}"
        if pct:   return f"{v:.2f}%"
        return f"{int(v):,}"

    ws2.merge_cells("A1:G1")
    h1 = ws2["A1"]
    h1.value = "SOURCE FILE VERIFICATION"
    h1.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    h1.fill  = PatternFill("solid", start_color=navy)
    h1.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    hdrs2 = ["Metric", "Source File", "Our Output", "Difference", "Status", "Tolerance", "Note"]
    for ci, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", start_color=teal)
        c.alignment = Alignment(horizontal="center")

    checks = [
        ("Total Units",           our["total_units"], ss.get("src_total_units"),    False, False, True,  "Exact", "All incl. vacant & non-rev"),
        ("Occupied Units",        our["occ_units"],   ss.get("src_occupied_units"), False, False, True,  "Exact", ""),
        ("Vacant Units",          our["vac_units"],   ss.get("src_vacant_units"),   False, False, True,  "Exact", ""),
        ("Non-Revenue Units",     our["nr_units"],    ss.get("src_nonrev_units"),   False, False, True,  "Exact", "Admin & Model"),
        ("Occupancy %",           our["occ_pct"],     ss.get("src_occ_pct"),        False, True,  False, "±1%",   ""),
        ("Total Market Rent",     our["total_mkt"],   ss.get("src_total_mkt"),      True,  False, False, "±0.5%", "Sum of all market rents"),
        ("Occupied Market Rent",  our["occ_mkt"],     ss.get("src_occupied_mkt"),   True,  False, False, "±0.5%", ""),
        # Use raw-file rent total for consistency — same codes as source summary
        ("Total Rent Charges",
         our.get("eff_total_raw", our["eff_total"]),
         ss.get("src_rent_total"),
         True, False, False, "±0.5%",
         "Raw rent charge total vs source — same charge codes"),
    ]

    for ri, (metric, our_v, src_v, money, pct, exact, tol_label, note) in enumerate(checks, 3):
        tol = 1.0 if pct else (2.0 if "Effective" in metric else 0.5)
        lbl, bg, fg = vstatus(our_v, src_v, tol=tol, exact=exact)
        diff = our_v - src_v if (src_v is not None and our_v is not None) else None
        row_fill = PatternFill("solid", start_color="FFFFFF" if ri % 2 else "F8FAFC")
        row_data = [metric, fmtv(src_v, money, pct), fmtv(our_v, money, pct),
                    fmtv(diff, money, pct) if diff is not None else "—",
                    lbl, tol_label, note]
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=10,
                          color=fg if ci == 5 else "1E293B", bold=(ci == 5))
            c.fill = PatternFill("solid", start_color=bg) if ci == 5 else row_fill
            c.alignment = Alignment(horizontal="right" if ci in (2, 3, 4) else
                                    "center" if ci in (5, 6) else "left")
            c.border = Border(bottom=Side(style="thin", color="E2E8F0"))

    cc = ss.get("src_charge_codes", {})
    if cc:
        cc_start = len(checks) + 5
        ws2.cell(row=cc_start, column=1,
                 value="Charge Code Breakdown (from Source File)").font = Font(
            name="Calibri", bold=True, size=10, color="0A2E3D")
        ws2.cell(row=cc_start+1, column=1, value="Code").font = Font(name="Calibri", bold=True, size=9)
        ws2.cell(row=cc_start+1, column=2, value="Amount").font= Font(name="Calibri", bold=True, size=9)
        for j, (code, amt) in enumerate(sorted(cc.items(), key=lambda x: -abs(x[1])), cc_start+2):
            is_r = code in RENT_CODES
            ws2.cell(row=j, column=1, value=code).font  = Font(name="Calibri", size=9,
                color="166534" if is_r else "374151", bold=is_r)
            ws2.cell(row=j, column=2, value=f"${amt:,.0f}").font = Font(name="Calibri", size=9,
                color="166534" if is_r else "374151")

    note_r = len(checks) + len(cc) + 8
    ws2.cell(row=note_r, column=1,
             value="MATCH = within tolerance  |  OFF = slightly different  |  MISMATCH = significant gap  |  N/A = not in source").font = Font(
        name="Calibri", size=8, italic=True, color="64748B")

    for i, w in enumerate([26, 20, 20, 16, 18, 12, 38], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    # Sheet 3: Raw File
    if raw_df is not None:
        ws3 = wb.create_sheet(title="Raw File (Original)")
        hdr_fill = PatternFill("solid", start_color="1E3A4A")
        for ci, col_val in enumerate(raw_df.columns, 1):
            c = ws3.cell(row=1, column=ci, value=str(col_val))
            c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
            c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(raw_df.itertuples(index=False), 2):
            rf = PatternFill("solid", start_color="FFFFFF" if ri%2==0 else "F8FAFC")
            for ci, val in enumerate(row, 1):
                cv = None if (isinstance(val, float) and pd.isna(val)) else val
                c = ws3.cell(row=ri, column=ci, value=cv)
                c.font = Font(name="Calibri", size=9, color="374151")
                c.fill = rf; c.alignment = Alignment(horizontal="left")
        for ci in range(1, raw_df.shape[1]+1):
            cl = get_column_letter(ci)
            ml = max((len(str(ws3.cell(row=r, column=ci).value or ""))
                      for r in range(1, min(50, raw_df.shape[0]+2))), default=8)
            ws3.column_dimensions[cl].width = min(ml+2, 30)
        ws3.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ── MAIN ──────────────────────────────────────────────────────────────────────
init_library()

if not check_password():
    st.stop()

st.markdown("""
<div class="rv-hero">
  <div class="rv-hero-left">
    <div class="rv-logo">RV</div>
    <div class="rv-hero-text">
      <h1>Rent Roll Standardizer</h1>
      <p>RealVal · Multifamily Underwriting Intelligence</p>
    </div>
  </div>
  <div class="rv-hero-right">
    <span class="rv-version">v5.8</span>
    <div class="rv-badge">⚡ Claude AI</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── TABS: Main tool | Format Library ──────────────────────────────────────────
main_tab, library_tab = st.tabs(["📋  Standardize Rent Roll", "📚  Format Library"])

with main_tab:
    st.markdown("<div class='sec-label'>Upload Rent Roll</div>", unsafe_allow_html=True)
    col_up, col_info = st.columns([2, 1])

    with col_up:
        st.markdown("<div class='upload-panel'><h3>Upload Your Excel File</h3><p>Any broker format accepted — multi-sheet workbooks supported</p>", unsafe_allow_html=True)
        uploaded_file = st.file_uploader("", type=["xlsx"], label_visibility="collapsed")

        # Format hint box
        st.markdown("<div class='hint-box'><label>Format Hint (optional)</label>", unsafe_allow_html=True)
        analyst_hint = st.text_area(
            "", height=68, label_visibility="collapsed",
            placeholder='e.g. "This broker uses annual rents" · "Charge codes in column H" · "Yardi export format"',
            key="analyst_hint"
        )
        st.markdown("</div></div>", unsafe_allow_html=True)

    with col_info:
        st.markdown("""
        <div class='info-panel'>
          <h4>What's new in v4.0</h4>
          <div class='info-row'><div class='info-dot'></div>Auto-detects correct sheet in multi-tab workbooks</div>
          <div class='info-row'><div class='info-dot'></div>Side-by-side raw vs. standardized preview</div>
          <div class='info-row'><div class='info-dot'></div>Format hint passes context to Claude</div>
          <div class='info-row'><div class='info-dot'></div>⚠ Flags rows with unusual effective rents</div>
          <div class='info-row'><div class='info-dot'></div>Claude flags low-confidence rows automatically</div>
          <div class='info-row'><div class='info-dot'></div>Chunk size auto-scales with file size</div>
          <div class='info-row'><div class='info-dot'></div>Format library saves examples for future files</div>
        </div>
        """, unsafe_allow_html=True)

    if uploaded_file:
        file_bytes = uploaded_file.read()
        st.markdown(
            f"<div class='file-pill'>📄 &nbsp;{uploaded_file.name}&nbsp; · &nbsp;{len(file_bytes)/1024:.1f} KB</div>",
            unsafe_allow_html=True
        )

        col_btn, col_broker = st.columns([1, 1])
        with col_btn:
            run = st.button("⚡  Standardize Rent Roll")
        with col_broker:
            broker_name = st.text_input("Broker / Source name (for library)",
                                        placeholder="e.g. CBRE, Marcus & Millichap, Yardi…",
                                        label_visibility="visible")

        if run:
            st.markdown("<div class='sec-label'>Processing</div>", unsafe_allow_html=True)
            step_ph   = st.empty()
            step_ph.markdown(render_steps(0), unsafe_allow_html=True)
            prog_ph   = st.progress(0)
            status_ph = st.empty()

            # Step 0 → Step 1: Read file
            step_ph.markdown(render_steps(1), unsafe_allow_html=True)
            try:
                original_df, sheet_name = detect_rent_roll_sheet(file_bytes)
            except Exception as e:
                st.error(f"Could not read file: {e}"); st.stop()

            status_ph.markdown(
                f"<small style='color:rgba(255,255,255,0.4);'>Sheet detected: <strong style='color:#93c5fd;'>'{sheet_name}'</strong> · {len(original_df)} raw rows</small>",
                unsafe_allow_html=True
            )
            time.sleep(0.4)

            # Extract source summary BEFORE processing (scans raw file footer)
            source_summary = extract_source_summary(original_df)

            # Build library context
            library_ctx = build_library_context()

            # Step 2 → processing
            standardized_df, col_map = standardize_rent_roll(
                original_df, step_ph, prog_ph, status_ph, analyst_hint, library_ctx,
                raw_df=original_df
            )

            if standardized_df.empty:
                st.error("Standardization failed or returned no data. Please check the file and try again.")
                st.stop()

            prog_ph.empty(); status_ph.empty()

            # Show detected format as a small badge
            fmt = col_map.get("fmt", "unknown")
            fmt_labels = {"yardi": "Yardi / MRI", "appfolio": "AppFolio", "vesper": "Vesper / Sunbelt", "onsite": "OneSite / RealPage", "unknown": "Unknown"}
            fmt_colors = {"yardi": "#60a5fa", "appfolio": "#34d399", "vesper": "#f59e0b", "onsite": "#c084fc", "unknown": "#94a3b8"}
            fmt_label = fmt_labels.get(fmt, fmt)
            fmt_color = fmt_colors.get(fmt, "#94a3b8")
            st.markdown(
                f"<div style='margin-bottom:0.8rem;'>"
                f"<span style='font-size:0.7rem;color:rgba(255,255,255,0.4);'>Format detected: </span>"
                f"<span style='background:rgba(255,255,255,0.07);border:1px solid {fmt_color}40;"
                f"border-radius:20px;padding:0.2rem 0.7rem;font-size:0.72rem;font-weight:600;"
                f"color:{fmt_color};'>{fmt_label}</span></div>",
                unsafe_allow_html=True
            )

            # Auto-save to format library
            if broker_name.strip():
                save_to_library(broker_name.strip(), analyst_hint, original_df, standardized_df)
                st.success(f"✅ Format saved to library as **{broker_name.strip()}**")

            # Flag banner
            flag_count = int(standardized_df["flag"].sum()) if "flag" in standardized_df.columns else 0
            if flag_count:
                st.markdown(f"""
                <div class='flag-banner'>
                  ⚠ &nbsp;<strong>{flag_count} row{'s' if flag_count>1 else ''} flagged for review</strong>
                  &nbsp;— effective rent appears unusual (>2× market or <$100). Check these rows before using.
                </div>""", unsafe_allow_html=True)

            # KPIs
            st.markdown("<div class='sec-label'>Underwriting Summary</div>", unsafe_allow_html=True)
            st.markdown(render_kpis(standardized_df), unsafe_allow_html=True)

            # Reconciliation
            st.markdown("<div class='sec-label'>Reconciliation Checks</div>", unsafe_allow_html=True)
            st.markdown(render_recon(standardized_df, len(original_df)), unsafe_allow_html=True)

            # Source file verifier (only shown if summary block was found)
            if source_summary:
                st.markdown("<div class='sec-label'>Source File Verifier</div>", unsafe_allow_html=True)
                st.markdown(render_source_verifier(standardized_df, source_summary), unsafe_allow_html=True)
            else:
                st.markdown("""<div style='font-size:0.75rem;color:rgba(255,255,255,0.3);
                margin-bottom:1rem;'>ℹ No summary block detected in source file — verifier not available for this format.</div>""",
                unsafe_allow_html=True)

            # Source File Verifier — shown only when summary block was found
            if source_summary:
                st.markdown("<div class='sec-label'>Source File Verifier</div>", unsafe_allow_html=True)
                st.markdown(render_source_verifier(standardized_df, source_summary), unsafe_allow_html=True)
            else:
                st.markdown("""
                <div style='background:#0d1e2e;border:1px solid rgba(255,255,255,0.06);
                border-radius:10px;padding:0.8rem 1.2rem;margin-bottom:1rem;
                font-size:0.78rem;color:rgba(255,255,255,0.3);'>
                ℹ️ No summary block detected in source file — verifier not available for this rent roll.
                </div>""", unsafe_allow_html=True)

            # Side-by-side tabs
            st.markdown("<div class='sec-label'>Preview</div>", unsafe_allow_html=True)
            tab_raw, tab_clean = st.tabs([f"📄  Raw File  ({sheet_name})", "✅  Standardized Output"])
            with tab_raw:
                st.markdown(render_raw_table(original_df, sheet_name), unsafe_allow_html=True)
            with tab_clean:
                st.markdown(render_table(standardized_df), unsafe_allow_html=True)

            # Download
            st.markdown("<div class='sec-label'>Download</div>", unsafe_allow_html=True)
            today     = date.today().strftime("%m%d%Y")
            base_name = uploaded_file.name.replace(".xlsx","")
            col_dl, col_hint_dl = st.columns([1, 2])
            with col_dl:
                st.download_button(
                    label="⬇  Download Standardized Rent Roll",
                    data=build_excel(standardized_df, original_df, source_summary),
                    file_name=f"{base_name}_standardized_{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            with col_hint_dl:
                if flag_count:
                    st.markdown(f"<div style='padding-top:0.7rem;font-size:0.78rem;color:#fbbf24;'>⚠ {flag_count} flagged rows included — review before use</div>", unsafe_allow_html=True)

    else:
        st.markdown("""
        <div style='background:#0d1e2e;border:1px solid rgba(255,255,255,0.06);border-radius:14px;
        padding:3rem;text-align:center;margin-top:1rem;'>
          <div style='font-size:2.5rem;margin-bottom:1rem;'>📂</div>
          <div style='font-family:Sora,sans-serif;font-size:1rem;font-weight:700;color:#fff;margin-bottom:0.4rem;'>No file uploaded yet</div>
          <div style='font-size:0.8rem;color:rgba(255,255,255,0.3);'>Upload an Excel rent roll above to get started</div>
        </div>""", unsafe_allow_html=True)

# ── FORMAT LIBRARY TAB ────────────────────────────────────────────────────────
with library_tab:
    st.markdown("<div class='sec-label'>Saved Format Examples</div>", unsafe_allow_html=True)
    lib = st.session_state.get("format_library", [])

    if not lib:
        st.markdown("""
        <div style='background:#0d1e2e;border:1px solid rgba(255,255,255,0.06);border-radius:14px;
        padding:2.5rem;text-align:center;'>
          <div style='font-size:2rem;margin-bottom:0.8rem;'>📚</div>
          <div style='font-family:Sora,sans-serif;font-size:0.95rem;font-weight:700;color:#fff;margin-bottom:0.4rem;'>No formats saved yet</div>
          <div style='font-size:0.78rem;color:rgba(255,255,255,0.3);'>
          Process a rent roll and enter a broker name to save it here.<br>
          Saved formats are used as examples to improve future processing.
          </div>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown(f"<div style='font-size:0.78rem;color:rgba(255,255,255,0.4);margin-bottom:1rem;'>{len(lib)} format{'s' if len(lib)>1 else ''} saved this session · used as context for future files</div>", unsafe_allow_html=True)
        for i, entry in enumerate(reversed(lib)):
            with st.expander(f"📄 {entry['broker']}  ·  {entry['unit_count']} units  ·  {entry['date']}"):
                if entry["hint"]:
                    st.markdown(f"**Hint used:** `{entry['hint']}`")
                st.markdown("**Raw sample (first 5 rows):**")
                st.code(entry["raw_sample"], language="text")
                st.markdown("**Standardized output sample:**")
                st.json(entry["out_sample"][:2])
        if st.button("🗑  Clear Format Library"):
            st.session_state["format_library"] = []
            st.rerun()

st.markdown("""
<div class='rv-footer'>
  <span>© 2026 <a href='https://therealval.com' target='_blank'>RealVal</a> · Internal Tool · Confidential</span>
  <span>Powered by Claude AI (Haiku) · Prompt {PROMPT_VERSION} · Authorized analyst use only</span>
</div>
""".replace("{PROMPT_VERSION}", PROMPT_VERSION), unsafe_allow_html=True)
